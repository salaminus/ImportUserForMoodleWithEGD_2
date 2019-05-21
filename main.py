###########################################
#       Экспорт ФИО из БД ЭЖД в Moodle    #
#               version 0.1               #
###########################################


import openpyxl
import transliterate


kItem = 0                                       #кол-во записей
listDataItems = []                              #список записей для экспорта в csv


#TODO: Открытие файла Excel
nameFileExcel = '10a-1'
nameFullFileExcel = nameFileExcel + '.xlsx'
wb = openpyxl.load_workbook(nameFullFileExcel)      #открытие книги Excel
listSheets = wb.sheetnames                      #получение списка листов книги Excel
sheets1 = wb[listSheets[0]]                     #берем первый лист книги Excel
# print(sheets1.cell(row=3,column=2).value)     #получение содержимого ячейки

def printInfoToConsole(rusLastName, rusFirstName, lastName, e_mail):
    """печать в консоль данных"""
    print(rusLastName + ' ' + rusFirstName + '; login: ' +
          lastName + '; password: ' + lastName +
          '; e-mail: ' + e_mail)

def translitNames(fio):
    """получение транслита фамилии и имени для логина, пароля и e-mail"""
    lastName = transliterate.translit(fio[0], reversed=True)
    lastName = lastName.replace("'", "")
    lastName = lastName.swapcase()[0] + lastName[1:]
    firstName = transliterate.translit(fio[1], reversed=True)
    firstName = firstName.replace("'", "")
    firstName = firstName.swapcase()[0] + firstName[1:]
    return lastName, firstName

def e_mailGet(lastName, firstName):
    """получение e-mail"""
    e_mail = lastName + '.' + firstName + '@test.ru'
    return e_mail

#TODO: Записать Ф и И в новый лист Excel и сохранить в csv
def writeDataToFileCSV(listDataItems):
    """запись данных в файл"""
    nameResultFile = nameFileExcel + '.csv'
    with open(nameResultFile,'w', encoding='WINDOWS-1251') as f:
        f.write('username' + ';' +
                'password' + ';' +
                'firstname' + ';' +
                'lastname' + ';' +
                'email' + '\n')
        for item in listDataItems:
            f.write(item['username'] + ';' +
                item['password'] + ';' +
                item['firstname'] + ';' +
                item['lastname'] + ';' +
                item['email'] + '\n')

#TODO: Записать данные в список listDataItem
def addToListDataItem(rusLastName, rusFirstName, lastName, e_mail):
    """создание словаря из rusLastName, rusFirstName, lastName, e_mail
       и запись в список словарей listDataItems"""
    dictDataItem = {}
    dictDataItem['username'] = lastName
    dictDataItem['password'] = lastName
    dictDataItem['firstname'] = rusFirstName
    dictDataItem['lastname'] = rusLastName
    dictDataItem['email'] = e_mail
    listDataItems.append(dictDataItem)

for i in range(3,34):
    # TODO: Получение столбца с ФИО
    valueCurrentCell = sheets1.cell(row=i,column=2).value
    if str(type(valueCurrentCell)) != "<class 'NoneType'>":
        try:
            fio = valueCurrentCell.split()
            # TODO: Разбить ФИО на фамилию и имя, удалить отчество
            rusLastName = fio[0]
            rusFirstName = fio[1]
            lastName = translitNames(fio)[0]
            firstName = translitNames(fio)[1]
            e_mail = e_mailGet(lastName, firstName)
            printInfoToConsole(rusLastName, rusFirstName, lastName, e_mail)
            addToListDataItem(rusLastName, rusFirstName, lastName, e_mail)
            if e_mail:
                kItem = kItem + 1

        except:
            pass
print('Total: ' + str(kItem))
writeDataToFileCSV(listDataItems)
print(listDataItems)
